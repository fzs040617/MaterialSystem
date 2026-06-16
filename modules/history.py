import streamlit as st
import pandas as pd
import datetime
import io
import json
import time
import uuid
import threading
import base64
from openpyxl.drawing.image import Image as xlImage
from openpyxl.utils import get_column_letter
from excel_engines import generate_rw_purchase_contract_excel

# 导入自定义模块
from database import get_db_conn, load_data
from business_logic import (
    undo_order, undo_other_material_order, 
    undo_inbound, undo_consumption, undo_purchase_order
)
from sync_history_to_feishu import (
    delete_feishu_purchase_order,
    migrate_purchase_orders_to_feishu,
    delete_feishu_other_material,
    delete_feishu_bag_order,
    delete_feishu_accessory_order,
    delete_feishu_inbound_order,
    delete_feishu_garment_consumption,
    delete_feishu_crossborder_order   
)

@st.dialog("✏️ 编辑采购合同", width="large")
def edit_purchase_order(po_id, role, user_name):
    """
    采购合同编辑逻辑 - 终极稳健版
    解决：1. 缓存不刷新 2. 组件记忆冲突 3. 数据类型丢失
    """
    if role != 'admin':
        st.error("无权限")
        return

    po_id = int(po_id)
    # 为每次弹窗生成一个绝对唯一的生命周期 ID，彻底斩断组件记忆
    if f"edit_token_{po_id}" not in st.session_state:
        st.session_state[f"edit_token_{po_id}"] = str(uuid.uuid4())[:8]
    
    token = st.session_state[f"edit_token_{po_id}"]
    required_cols = ['图片', '物料编号', '物料名称', '材质', '颜色', '尺寸', '收货标准', '数量', '单位', '单价', '货期', '备注']
    # 1. 【核心：强制实时读】不使用 load_data 缓存函数，直接建立新连接抓取
    conn = get_db_conn()
    cursor = conn.cursor()
    cursor.execute("SELECT id, contract_no, factory_name, remark FROM purchase_orders WHERE id=%s", (po_id,))
    row = cursor.fetchone()
    conn.close()

    if not row:
        st.error("❌ 找不到该合同记录")
        return
    
    _, contract_no, factory_name, remark_json = row

    # 2. 解析 JSON 结构
    try:
        orig_rem = json.loads(remark_json) if remark_json else {}
        items = orig_rem.get('items', orig_rem) if isinstance(orig_rem, dict) else orig_rem
        if not isinstance(items, list): items = []
    except:
        items = []
        orig_rem = {}

    # 3. 准备表格数据并托管到 Session State 以支持外部修改
    state_key = f"df_items_{po_id}_{token}"
    if state_key not in st.session_state:
        df_items = pd.DataFrame(items)
        for col in required_cols:
            if col not in df_items.columns:
                df_items[col] = '' if col not in ['数量', '单价'] else 0
        if '货期' in df_items.columns:
            df_items['货期'] = pd.to_datetime(df_items['货期'], errors='coerce').dt.date
        st.session_state[state_key] = df_items

    st.markdown(f"**📄 合同编号：** `{contract_no}`")

    # 4. 使用表单锁定数据流
    with st.form(key=f"po_edit_form_{po_id}_{token}"):
        new_fac = st.text_input("🏭 乙方工厂", value=factory_name)
        st.info("💡 修改完毕后，请点击下方【覆盖保存】按钮。系统将同步更新数据库与 Excel 附件。")
        
        # 🌟 新增：编辑弹窗中的图片上传助手
        with st.expander("🖼️ 图片上传助手 (针对特定行插入或更换本地图片)"):
            c_r, c_u = st.columns([1, 3])
            with c_r:
                tgt_row = st.number_input("🎯 目标行号", min_value=1, max_value=len(st.session_state[state_key]), step=1)
            with c_u:
                up_img = st.file_uploader("📤 选择新图片", type=['png','jpg','jpeg'], key=f"up_{po_id}_{token}")
                if up_img:
                    b64 = base64.b64encode(up_img.getvalue()).decode()
                    b64_str = f"data:{up_img.type};base64,{b64}"
                    st.session_state[state_key].iat[tgt_row-1, st.session_state[state_key].columns.get_loc('图片')] = b64_str
                    st.success("✅ 图片已成功替换！")

        edited_df = st.data_editor(
            st.session_state[state_key][required_cols],
            column_config={
                "图片": st.column_config.ImageColumn("🖼️ 图片 (可粘贴链接)"),
                "单价": st.column_config.NumberColumn("单价", format="%.4f"),
                "数量": st.column_config.NumberColumn("数量", format="%d"),
                "货期": st.column_config.DateColumn("货期"),
            },
            # ...(后续保持不变)
            hide_index=True,
            use_container_width=True,
            num_rows="dynamic",
            key=f"editor_{po_id}_{token}"
        )
        
        submit = st.form_submit_button("💾 确认并覆盖保存", type="primary", use_container_width=True)

    if submit:
        try:
            # 5. 数据清洗：强制将所有 Numpy 类型转回纯 Python 类型，防止 JSON 报错
            edited_df['数量'] = pd.to_numeric(edited_df['数量'], errors='coerce').fillna(0).astype(int)
            edited_df['单价'] = pd.to_numeric(edited_df['单价'], errors='coerce').fillna(0.0).astype(float)
            edited_df = edited_df.fillna("")

            for idx in edited_df.index:
                val = edited_df.loc[idx, '货期']
                edited_df.loc[idx, '货期'] = val.strftime('%Y-%m-%d') if hasattr(val, 'strftime') else str(val)

            new_items = json.loads(edited_df.to_json(orient='records', force_ascii=False))
            
            # 重新封装 JSON
            new_rem_obj = orig_rem if isinstance(orig_rem, dict) else {}
            new_rem_obj['items'] = new_items
            new_rem_obj['contract_no'] = contract_no
            final_remark_str = json.dumps(new_rem_obj, ensure_ascii=False)

            # 6. 生成新 Excel
            new_excel = generate_rw_purchase_contract_excel(contract_no, new_fac, new_items)

            # 7. 【关键：事务写入】
            conn = get_db_conn()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE purchase_orders 
                SET factory_name=%s, remark=%s, excel_data=%s 
                WHERE id=%s
            """, (new_fac, final_remark_str, new_excel, po_id))
            conn.commit() # 👈 必须手动提交
            conn.close()

            # ==========================================
            # 🌟 新增：触发飞书同步更新（先删后补策略）
            # ==========================================
            def sync_edit_to_feishu(pid):
                try:
                    # 1. 先把飞书里旧的明细行全部干掉
                    delete_feishu_purchase_order(pid)
                    
                    time.sleep(1) # 缓冲1秒，等待飞书服务器消化删除指令
                    
                    # 2. 触发全量同步。因为旧的刚被删了，防重机制会判定它不存在
                    # 从而自动把你刚刚 UPDATE 的最新数据重新推送到飞书！
                    migrate_purchase_orders_to_feishu()
                except Exception as e:
                    print(f"编辑合同同步飞书失败: {e}")
                    
            # 异步执行，绝不卡顿前端页面
            threading.Thread(target=sync_edit_to_feishu, args=(po_id,)).start()
            # ==========================================

            # 8. 【核心修复：暴力刷缓存】
            st.cache_data.clear() # 👈 清除所有 load_data 的历史缓存
            if f"edit_token_{po_id}" in st.session_state:
                del st.session_state[f"edit_token_{po_id}"]
            
            st.success("✅ 修改成功！数据已实时同步至数据库与飞书多维表格。")
            time.sleep(1)
            st.rerun()

        except Exception as e:
            st.error(f"❌ 保存失败：{str(e)}")

def render_history(role, user_name):
    """渲染历史记录与回溯中心"""
    st.header("📜 历史记录与回溯")
    
    # 定义顶层标签页
    tab_po, tab_omat, tab_out, tab_acc, tab_in, tab_consume, tab_cross = st.tabs([
        "🛒 采购合同记录", "📦 其他物料记录", "📋 包装袋下单记录", 
        "🖨️ 辅料下单记录", "📥 入库记录", "👕 制衣厂消耗记录", "🌍 跨境物料记录"
    ])

    # ==========================================
    # 1. 采购合同历史记录 (Purchase Orders)
    # ==========================================
    with tab_po:
        st.subheader("🛒 采购合同历史记录")
        # 强制使用新连接直读数据库，打破 SQLAlchemy 缓存幻觉
        conn_po = get_db_conn()
        try:
            df_po = pd.read_sql_query("SELECT * FROM purchase_orders ORDER BY id DESC", conn_po)
        except Exception:
            df_po = pd.DataFrame()
        finally:
            conn_po.close()
        if df_po.empty:
            st.info("📭 暂无采购合同记录")
        else:
            # 筛选与导出逻辑
            st.markdown("##### 🔍 多维度筛选与明细导出")
            month_list_po = sorted(list(set([str(d)[:7] for d in df_po['create_time'] if pd.notna(d)])), reverse=True)
            fac_list_po = sorted(list(set([str(f) for f in df_po['factory_name'] if pd.notna(f)])))

            c_p1, c_p2 = st.columns(2)
            sel_po_months = c_p1.multiselect("📅 合同月份", month_list_po, key="po_exp_month")
            sel_po_facs = c_p2.multiselect("🏭 乙方工厂", fac_list_po, key="po_exp_fac")

            df_po_filtered = df_po.copy()
            if sel_po_months:
                df_po_filtered = df_po_filtered[df_po_filtered['create_time'].str[:7].isin(sel_po_months)]
            if sel_po_facs:
                df_po_filtered = df_po_filtered[df_po_filtered['factory_name'].isin(sel_po_facs)]

            # 单份合同下载与撤销
            # --- [新增：物料明细平铺展示逻辑] ---
            st.markdown("##### 📊 采购合同物料明细清单")
        
            all_details = []
            for _, row in df_po_filtered.iterrows():
                try:
                    raw_data = json.loads(row['remark']) if row['remark'] else []
                    if isinstance(raw_data, list):
                        items = raw_data
                    elif isinstance(raw_data, dict):
                        items = raw_data.get('items', [])
                    else:
                        items = []
                    for item in items:
                        all_details.append({
                            "合同编号": row['contract_no'],
                            "系统单号": f"PO-{row['id']:04d}",
                            "创建时间": row['create_time'],
                            "乙方工厂": row['factory_name'],
                            "操作人": row['operator'],
                            "图片": item.get('图片', ''),
                            "物料编号": item.get('物料编号', ''),
                            "物料名称": item.get('物料名称', ''),
                            "材质": item.get('材质', ''),
                            "尺寸": item.get('尺寸', ''),
                            "颜色": item.get('颜色', ''),
                            "收货标准": item.get('收货标准', ''),
                            "数量": item.get('数量', 0),
                            "单位": item.get('单位', 'Pcs'),
                            "单价(含税运)": item.get('单价', 0.0),
                            "总金额": round(float(item.get('数量', 0)) * float(item.get('单价', 0.0)), 3),
                            "货期": item.get('货期', ''),
                            "备注": item.get('备注', '')
                        })
                except Exception as e:
                    continue
        
            if all_details:
                df_display = pd.DataFrame(all_details)
                # 按照单号倒序排列
                df_display = df_display.sort_values("系统单号", ascending=False)
                
                # ---------- 显示优化：自动换行 + 合理列宽 ----------
                column_config = {
                    "合同编号": st.column_config.TextColumn("合同编号", width="medium"),
                    "系统单号": st.column_config.TextColumn("系统单号", width="small"),
                    "创建时间": st.column_config.TextColumn("创建时间", width="medium"),
                    "乙方工厂": st.column_config.TextColumn("乙方工厂", width="medium"),
                    "操作人": st.column_config.TextColumn("操作人", width="small"),
                    "物料编号": st.column_config.TextColumn("物料编号", width="small"),
                    "物料名称": st.column_config.TextColumn("物料名称", width="large"),
                    "材质": st.column_config.TextColumn("材质", width="large"),
                    "尺寸": st.column_config.TextColumn("尺寸", width="medium"),
                    "颜色": st.column_config.TextColumn("颜色", width="small"),
                    "收货标准": st.column_config.TextColumn("收货标准", width="large"),
                    "数量": st.column_config.NumberColumn("数量", format="%d", width="small"),
                    "单位": st.column_config.TextColumn("单位", width="small"),
                    "单价(含税运)": st.column_config.NumberColumn("单价(含税运)", format="%.4f", width="small"),
                    "总金额": st.column_config.NumberColumn("总金额", format="%.2f", width="medium"),
                    "货期": st.column_config.TextColumn("货期", width="small"),
                    "备注": st.column_config.TextColumn("备注", width="large")
                }
                st.dataframe(
                    df_display,
                    column_config=column_config,
                    use_container_width=True,
                    hide_index=True,
                    height=400  # 固定高度，内部滚动
                )
                
                # ---------- 导出优化：改为 Excel 并自动调整列宽 + 贴图 ----------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_display.to_excel(writer, index=False, sheet_name='采购明细')
                    worksheet = writer.sheets['采购明细']
                    
                    # 自动调整基础列宽
                    for column in worksheet.columns:
                        max_length = 0
                        col_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if cell.value:
                                    cell_len = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in str(cell.value))
                                    if cell_len > max_length:
                                        max_length = cell_len
                            except:
                                pass
                        adjusted_width = min(max(max_length + 2, 8), 50)
                        worksheet.column_dimensions[col_letter].width = adjusted_width
                    
                    # 🌟 核心：检测并渲染图片列
                    if "图片" in df_display.columns:
                        img_col_idx = df_display.columns.get_loc("图片") + 1
                        img_col_letter = get_column_letter(img_col_idx)
                        worksheet.column_dimensions[img_col_letter].width = 15 # 强制撑开图片列的宽度
                        
                        # 遍历数据行 (start=2 是因为跳过 Excel 第一行表头)
                        for r_idx, row_data in enumerate(df_display.itertuples(), start=2):
                            worksheet.row_dimensions[r_idx].height = 60 # 撑开行高适应图片
                            img_str = getattr(row_data, "图片", "")
                            
                            # 解析 Base64 并画入表格
                            if isinstance(img_str, str) and img_str.startswith('data:image'):
                                try:
                                    header, encoded = img_str.split(",", 1) if "," in img_str else ("", img_str)
                                    img_data = base64.b64decode(encoded)
                                    img = xlImage(io.BytesIO(img_data))
                                    img.width = 60
                                    img.height = 60
                                    worksheet.add_image(img, f"{img_col_letter}{r_idx}")
                                    worksheet.cell(row=r_idx, column=img_col_idx, value="") # 清空乱码文本
                                except Exception:
                                    worksheet.cell(row=r_idx, column=img_col_idx, value="图片损坏")
                            elif isinstance(img_str, str) and img_str.startswith('http'):
                                # 兼容网络链接
                                worksheet.cell(row=r_idx, column=img_col_idx, value=img_str)
                            else:
                                worksheet.cell(row=r_idx, column=img_col_idx, value="")
                    
                    # 文本自动换行
                    for row in worksheet.iter_rows():
                        for cell in row:
                            if cell.value and isinstance(cell.value, str):
                                cell.alignment = cell.alignment.copy(wrap_text=True)
                st.download_button(
                    "📥 导出当前筛选的明细表 (Excel 格式，已优化列宽)",
                    data=output.getvalue(),
                    file_name="采购明细导出.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("查无明细数据")
        
            st.divider()
            
            po_options = df_po_filtered['create_time'] + " | " + df_po_filtered['factory_name']
            sel_po_str = st.selectbox("选择历史记录：", po_options, key="sel_single_po")
            if sel_po_str:
                sel_row = df_po_filtered[po_options == sel_po_str].iloc[0]
    
                col_dl, col_edit, col_del = st.columns(3)
                with col_dl:
                    st.download_button("⬇️ 下载合同", sel_row['excel_data'], f"合同_{sel_row['id']}.xlsx")
                with col_edit:
                    if role == 'admin':
                        if st.button("✏️ 编辑合同", key=f"edit_po_{sel_row['id']}"):
                            st.session_state['editing_po_id'] = sel_row['id']
                            st.rerun()
                with col_del:
                    if (role == 'admin' or sel_row['operator'] == user_name):
                        if st.button(f"❌ 撤销并删除", type="secondary"):
                            ok, msg = undo_purchase_order(sel_row['id'])
                            if ok: st.success(msg); time.sleep(1); st.rerun()

    # ==========================================
    # 2. 其他物料下单记录 (Other Materials)
    # ==========================================
    with tab_omat:
        h_omat = load_data("other_material_history")
        if h_omat.empty: 
            st.info("📦 暂无其他物料下单数据")
        else:
            st.subheader("📦 其他物料发货历史")
            h_omat = h_omat.sort_values("id", ascending=False)
            st.dataframe(h_omat[['id','order_date','operator','material_display','total_quantity']], use_container_width=True, hide_index=True)
            st.divider()
            
            c_sel, c_view = st.columns([1,2])
            with c_sel: 
                omat_id = st.selectbox("👉 订单ID (查看/撤销)", h_omat['id'].tolist(), key="so_omat")
            if omat_id and (omat_id in h_omat['id'].values):
                row = h_omat[h_omat['id']==omat_id].iloc[0]
                can_undo = (role == 'admin') or (row.get('operator') == user_name)
                with c_view:
                    if can_undo:
                        # 在 render_history 的 tab_omat 标签页内：
                        if st.button(f"🗑️ 撤销该发货单 #{omat_id}", type="secondary"):
                            ok, msg = undo_other_material_order(omat_id)
                            if ok: 
                                # 🌟 埋点触发删除
                                threading.Thread(target=delete_feishu_other_material, args=(omat_id,)).start()
                                st.success(msg); time.sleep(1); st.rerun()
                            else: st.error(msg)
                    else: st.caption("🔒 无撤销权限")
                try:
                    d = json.loads(row['details'])
                    s = d.get('shipping',[])
                    st.markdown(f"**物料：** {row['material_display']}")
                    cart_list = []
                    for item in s:
                        cart_list.append({"发货工厂":item['src_factory'], "数量":item['qty'], "收货制衣厂":item['dst_garment']})
                    st.write("🚚 详情路线")
                    st.dataframe(pd.DataFrame(cart_list), use_container_width=True, hide_index=True)
                except: 
                    st.error("数据解析失败")

    # ==========================================
    # 3. 包装袋下单记录 (Packaging Bags)
    # ==========================================
    with tab_out:
        h = load_data("order_history")
        if h.empty: 
            st.info("📦 暂无包装袋下单数据")
        else:
            # ---------------------------------------------------------
            # [升级版功能]：按【月份】或源工厂维度筛选并合并导出
            # ---------------------------------------------------------
            # ---------------------------------------------------------
            # [终极版]：按【月份】和【源工厂】双重组合筛选并合并导出
            # ---------------------------------------------------------
            # ---------------------------------------------------------
            # [终极版]：按【月份】、【源工厂】、【商品名称】三重组合筛选并合并导出
            # ---------------------------------------------------------
            st.markdown("##### 📥 订单多维度组合筛选与合并导出")
            
            # 1. 提取月份、工厂、商品名称列表
            month_list = sorted(list(set([str(d)[:7] for d in h['order_date'] if pd.notna(d)])), reverse=True)
            if not month_list:
                month_list = [datetime.date.today().strftime("%Y-%m")]
                
            pack_facts_list = load_data("packaging_factories")['name'].tolist()
            
            # [新增] 提取历史记录里所有的商品名称，并去重排序
            prod_list = sorted(list(set([str(p).strip() for p in h['product_name'] if pd.notna(p) and str(p).strip() != ''])))
            
            
            # ==========================================
            # 🌟 修改：加载库存字典（包含工厂维度），用于精准匹配单价
            # ==========================================
            inv_df = load_data("inventory")
            price_map = {}
            if not inv_df.empty:
                for _, inv_row in inv_df.iterrows():
                    # 以 "工厂_名称_尺寸" 作为唯一联合键
                    fac = str(inv_row.get('factory_name', '')).strip()
                    b_n = str(inv_row.get('bag_name', '')).strip()
                    b_s = str(inv_row.get('bag_size', '')).strip()
                    key = f"{fac}_{b_n}_{b_s}"
                    try:
                        price_map[key] = float(inv_row.get('unit_price', 0.0))
                    except:
                        price_map[key] = 0.0
            # ==========================================

            # 2. 渲染三重多选框 (分成三列排版，留空代表不限制该条件)
            c_filt1, c_filt2, c_filt3 = st.columns(3)
            with c_filt1:
                sel_months = st.multiselect("📅 下单月份 (多选)", month_list, key="bag_exp_month")
            with c_filt2:
                sel_facs = st.multiselect("🏭 发货源工厂 (多选)", pack_facts_list, key="bag_exp_fac")
            with c_filt3:
                sel_prods = st.multiselect("📦 商品名称 (多选)", prod_list, key="bag_exp_prod")
                
            # 3. 实时计算筛选结果
            export_rows = []
            for _, row in h.iterrows():
                row_month = str(row['order_date'])[:7]
                row_prod = str(row['product_name']).strip()
                
                # 限制1：月份筛选 (不符合则跳过)
                if sel_months and row_month not in sel_months:
                    continue
                    
                # [新增] 限制2：商品名称筛选 (不符合则跳过)
                if sel_prods and row_prod not in sel_prods:
                    continue
                    
                try:
                    d = json.loads(row['details'])
                    shipping_list = d.get('shipping', [])
                    
                    for item in shipping_list:
                        src_factory = item.get('src_factory', '')
                        
                        # 限制3：工厂筛选 (因为明细里的工厂可能不同，所以放在深层判断)
                        if sel_facs and src_factory not in sel_facs:
                            continue
                            
                        # 过了三关，满足所有条件，装入导出列表
                        
                        # 🌟 修改：使用包含工厂的 key 进行三维精准匹配
                        bag_n = str(row['bag_name']).strip()
                        bag_s = str(row['bag_size']).strip()
                        # 注意：这里的 src_factory 是从循环里 item 字典提取出来的明细工厂
                        fac_str = str(src_factory).strip() 
                        
                        bag_key = f"{fac_str}_{bag_n}_{bag_s}"
                        u_price = price_map.get(bag_key, 0.0)
                        
                        qty = item.get('qty', 0)
                        total_price = u_price * qty
                        
                        export_rows.append({
                            '日期': row['order_date'],
                            '源工厂': src_factory,
                            '销售平台': row['platform'],
                            '商品名称': row['product_name'],
                            '数量': qty,
                            '单价': u_price,       # 👈 新增
                            '小计': total_price,   # 👈 新增
                            '发往制衣厂': item.get('dst_garment', ''),
                            '包装袋名称': row['bag_name'],
                            '尺寸': row['bag_size']
                        })
                except:
                    continue
                    
            # 4. 动态展示结果与下载按钮
            if not export_rows:
                st.info("💡 请在上方选择条件。当前条件下暂无可导出的发货明细。")
            else:
                st.success(f"✅ 当前条件下，共精准匹配到 **{len(export_rows)}** 条发货明细！")
                
                # 构建 Excel
                df_export = pd.DataFrame(export_rows)
                # 🌟 修改：加入单价和小计列排版
                cols_order = ['日期', '源工厂', '销售平台', '商品名称', '包装袋名称', '尺寸', '数量', '单价', '小计', '发往制衣厂']
                df_export = df_export[cols_order]
                
                out = io.BytesIO()
                with pd.ExcelWriter(out, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='下单汇总')
                    worksheet = writer.sheets['下单汇总']
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.5
                        worksheet.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50
                
                # 动态生成文件名，变得更聪明
                name_parts = ["包装袋汇总"]
                if sel_months: name_parts.append("&".join(sel_months))
                if sel_facs: name_parts.append("&".join(sel_facs))
                # 考虑到商品名可能很多，如果选了商品名，只取前两个拼接防止文件名太长报错
                if sel_prods: 
                    prod_str = "&".join(sel_prods[:2]) + ("等" if len(sel_prods)>2 else "")
                    name_parts.append(prod_str)
                    
                export_filename = "_".join(name_parts) + ".xlsx"
                
                st.download_button(
                    label=f"📥 立即下载 Excel 汇总表", 
                    data=out.getvalue(), 
                    file_name=export_filename, 
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True
                )
            
            # ---------------------------------------------------------
            # [升级版] 按时间段筛选并批量撤销
            # ---------------------------------------------------------
            st.markdown("##### 🗑️ 按时间段筛选并批量撤销")
            
            # 1. 日期区间选择器 (默认选中今天)
            date_range = st.date_input(
                "📅 选择下单时间段进行筛选", 
                value=(datetime.date.today(), datetime.date.today()), 
                key="bag_undo_date"
            )
            
            # 2. 根据日期区间过滤底层数据
            df_undo_view = h.copy()
            if len(date_range) == 2:
                start_d, end_d = date_range
                df_undo_view['dt_obj'] = pd.to_datetime(df_undo_view['order_date']).dt.date
                df_undo_view = df_undo_view[(df_undo_view['dt_obj'] >= start_d) & (df_undo_view['dt_obj'] <= end_d)]
            elif len(date_range) == 1:
                start_d = date_range[0]
                df_undo_view['dt_obj'] = pd.to_datetime(df_undo_view['order_date']).dt.date
                df_undo_view = df_undo_view[df_undo_view['dt_obj'] == start_d]
            
            df_undo_view = df_undo_view.sort_values("id", ascending=False)
            
            if df_undo_view.empty:
                st.info("💡 该时间段内没有包装袋的发货记录。")
            else:
                # 3. 渲染可打勾多选的数据表格
                df_undo_disp = df_undo_view[['id','order_date','operator','product_name','bag_name','bag_size','total_quantity']].copy()
                # 在最前面强行插入一列用于打勾
                df_undo_disp.insert(0, "✅ 勾选撤销", False)
                
                edited_undo = st.data_editor(
                    df_undo_disp,
                    column_config={
                        "✅ 勾选撤销": st.column_config.CheckboxColumn("选错可撤销", default=False),
                        "id": st.column_config.NumberColumn("单号", disabled=True)
                    },
                    # 禁用除打勾以外的所有列，防止用户在这里随意篡改历史数据
                    disabled=['order_date','operator','product_name','bag_name','bag_size','total_quantity'],
                    hide_index=True,
                    use_container_width=True,
                    key="batch_undo_editor_bag"
                )
                
                # 4. 提取被勾选的 ID 列表
                selected_ids = edited_undo[edited_undo["✅ 勾选撤销"] == True]['id'].tolist()
                
                # 5. 批量撤销的触发逻辑
                if selected_ids:
                    st.warning(f"⚠️ 警告：您已勾选 **{len(selected_ids)}** 条记录！点击下方按钮将把这些单据关联的库存全部退回。")
                    if st.button("🚨 确认执行批量撤销", type="primary", use_container_width=True):
                        success_count = 0
                        fail_count = 0
                        success_ids = []
                        
                        # 遍历执行撤销指令
                        for uid in selected_ids:
                            row_info = df_undo_view[df_undo_view['id'] == uid].iloc[0]
                            # 双重保险权限校验：管理员可以撤销所有，其他人只能撤销自己操作的
                            if role == 'admin' or row_info['operator'] == user_name:
                                ok, msg = undo_order(uid)
                                if ok: 
                                    success_count += 1
                                    success_ids.append(uid)
                                else: 
                                    fail_count += 1
                            else:
                                fail_count += 1 # 权限不足也算失败

                        if success_ids:
                            def batch_sync_delete(ids):
                                for sid in ids:
                                    try:
                                        delete_feishu_bag_order(sid)
                                    except Exception:
                                        pass
                            threading.Thread(target=batch_sync_delete, args=(success_ids,)).start()
                                
                        st.success(f"✅ 批量操作完成！成功撤销 {success_count} 条，失败/无权限 {fail_count} 条。")
                        time.sleep(1.5)
                        st.rerun()

    # ==========================================
    # 4. 辅料下单记录 (Accessories)
    # ==========================================
    with tab_acc:
        st.subheader("🖨️ 辅料开单历史")
        conn = get_db_conn()
        try:
            df_acc = pd.read_sql_query("SELECT * FROM accessory_history ORDER BY id DESC", conn)
        except:
            df_acc = pd.DataFrame()
        
        if df_acc.empty:
            st.info("📭 暂无辅料开单记录")
        else:
            # --- 3.1 批量汇总导出区块 ---
            # --- 3.1 多维度组合筛选与批量导出区块 ---
            st.markdown("##### 📥 多维度组合筛选与批量导出")
            
            # 1. 提取筛选项列表并去重排序 (将辅料款式替换为了产品名称)
            month_list = sorted(list(set([str(d)[:7] for d in df_acc['order_date'] if pd.notna(d)])), reverse=True)
            fac_list = sorted(list(set([str(f).strip() for f in df_acc['factory_name'] if pd.notna(f) and str(f).strip() != ''])))
            prod_list = sorted(list(set([str(p).strip() for p in df_acc['product_name'] if pd.notna(p) and str(p).strip() != ''])))

            # 2. 渲染横向的三重筛选器
            c_filt1, c_filt2, c_filt3 = st.columns(3)
            with c_filt1:
                sel_months = st.multiselect("📅 下单月份 (多选)", month_list, key="acc_exp_month")
            with c_filt2:
                sel_facs = st.multiselect("🏭 收货制衣厂 (多选)", fac_list, key="acc_exp_fac")
            with c_filt3:
                sel_prods = st.multiselect("📦 产品名称 (多选)", prod_list, key="acc_exp_prod")

            # 3. 实时过滤数据
            df_acc_filtered = df_acc.copy()
            if sel_months:
                df_acc_filtered = df_acc_filtered[df_acc_filtered['order_date'].astype(str).str[:7].isin(sel_months)]
            if sel_facs:
                df_acc_filtered = df_acc_filtered[df_acc_filtered['factory_name'].isin(sel_facs)]
            if sel_prods:
                df_acc_filtered = df_acc_filtered[df_acc_filtered['product_name'].isin(sel_prods)]

            # 4. 渲染可打勾的表格与动态导出
            if df_acc_filtered.empty:
                st.info("💡 当前筛选条件下暂无可导出的辅料记录。")
            else:
                st.success(f"✅ 筛选成功：共匹配到 **{len(df_acc_filtered)}** 条辅料明细。请在下方表格勾选需要导出的行（点击表头复选框可一键全选）。")
                
                df_acc_disp = df_acc_filtered.copy()
                df_acc_disp.insert(0, "✅ 选择汇总", False)
                
                col_map = {
                    'order_date': '下单日期', 'item_no': '货号', 'product_name': '产品名称',
                    'acc_style': '辅料款式', 'total_qty': '下单数量', 'factory_name': '收货制衣厂',
                    'internal_code': '内部码', 'material_info': '成分资料'
                }
                
                # 自动识别存在的列
                actual_cols = ["✅ 选择汇总"] + [c for c in col_map.keys() if c in df_acc_disp.columns]
                
                edited_df = st.data_editor(
                    df_acc_disp[actual_cols].rename(columns=col_map),
                    hide_index=True,
                    use_container_width=True,
                    column_config={"✅ 选择汇总": st.column_config.CheckboxColumn(width="small")},
                    key="acc_summary_editor"
                )

                selected_indices = edited_df[edited_df["✅ 选择汇总"] == True].index
                if not selected_indices.empty:
                    # 智能拼接导出文件名 (防超长处理)
                    name_parts = ["辅料汇总"]
                    if sel_months: name_parts.append("&".join(sel_months))
                    if sel_facs: name_parts.append("&".join(sel_facs))
                    if sel_prods: 
                        prod_str = "&".join(sel_prods[:2]) + ("等" if len(sel_prods)>2 else "")
                        name_parts.append(prod_str)
                        
                    export_filename = "_".join(name_parts) + ".xlsx"

                    if st.button(f"🚀 生成并导出这 {len(selected_indices)} 条记录的汇总表", type="primary"):
                        export_df = edited_df.loc[selected_indices].drop(columns=["✅ 选择汇总"])
                        output = io.BytesIO()
                        with pd.ExcelWriter(output, engine='openpyxl') as writer:
                            export_df.to_excel(writer, index=False, sheet_name='辅料下单汇总')
                            
                            # 自动列宽排版
                            worksheet = writer.sheets['辅料下单汇总']
                            for column in worksheet.columns:
                                max_length = 0
                                column_letter = get_column_letter(column[0].column)
                                for cell in column:
                                    try:
                                        if len(str(cell.value)) > max_length:
                                            max_length = len(str(cell.value))
                                    except: pass
                                adjusted_width = (max_length + 2) * 1.5
                                worksheet.column_dimensions[column_letter].width = adjusted_width if adjusted_width < 50 else 50

                        st.download_button(
                            label="📥 点击下载 Excel 汇总表", 
                            data=output.getvalue(), 
                            file_name=export_filename, 
                            type="primary"
                        )

            # --- 3.2 单份文件重下与撤销 ---
            st.markdown("##### 📂 单份文件重下与撤销")
            record_options = df_acc['create_time'] + " | " + df_acc['factory_name']
            sel_record_str = st.selectbox("请选择要操作的历史记录：", record_options, key="sel_single_acc")
            
            if sel_record_str:
                sel_idx = record_options[record_options == sel_record_str].index[0]
                sel_id = df_acc.iloc[sel_idx]['id']
                sel_filename = df_acc.iloc[sel_idx]['file_name']
                
                c_dl, c_del = st.columns(2)
                with c_dl:
                    cursor = conn.cursor()
                    cursor.execute("SELECT excel_data FROM accessory_history WHERE id=%s", (int(sel_id),))
                    file_blob = cursor.fetchone()[0]
                    st.download_button(f"⬇️ 重新下载: {sel_filename}", file_blob, file_name=sel_filename, type="primary", use_container_width=True)
                with c_del:
                    if st.button(f"❌ 撤销记录 #{sel_id}", type="secondary", use_container_width=True):
                        cursor = conn.cursor()
                        cursor.execute("DELETE FROM accessory_history WHERE id=%s", (int(sel_id),))
                        conn.commit()
                        # 异步删除飞书中的对应记录
                        threading.Thread(target=delete_feishu_accessory_order, args=(sel_id,)).start()
                        st.success("✅ 记录已删除"); time.sleep(1); st.rerun()
                        
        conn.close()

    # ==========================================
    # 5. 入库记录与消耗记录
    # ==========================================
    with tab_in:
        ih = load_data("inbound_history")
        if ih.empty: 
            st.info("📥 暂无入库数据")
        else:
            ih = ih.sort_values("id", ascending=False)
            st.dataframe(ih, use_container_width=True, hide_index=True)
            st.divider()
            c1, c2 = st.columns([1,2])
            with c1: 
                iid = st.selectbox("👉 入库ID (撤销)", ih['id'].tolist(), key="si_inbound")
            if iid:
                irow = ih[ih['id']==iid].iloc[0]
                can = (role == 'admin') or (irow.get('operator') == user_name)
                with c2:
                    st.write(f"详情: {irow['factory_name']} +{irow['quantity']}")
                    if can:
                        if st.button(f"🗑️ 撤销入库 #{iid}", type="secondary"):
                            ok, msg = undo_inbound(iid)
                            if ok:
                                # 异步删除飞书中的对应记录
                                threading.Thread(target=delete_feishu_inbound_order, args=(iid,)).start()
                                st.success(msg); time.sleep(1); st.rerun()
                            else:
                                st.error(msg)
                    else: st.caption("🔒 无权限")
    
    with tab_consume:
        gc = load_data("garment_consumption")
        if gc.empty:
            st.info("📦 暂无制衣厂消耗数据")
        else:
            st.markdown("##### 🔍 按订单号检索与导出")
            # 1. 检索输入框
            search_order = st.text_input("输入关联生产订单号 / 款号进行检索 (留空则显示全部)", key="search_consume_order")
            
            # 2. 动态过滤数据
            if search_order.strip():
                # 模糊匹配订单号
                df_filtered = gc[gc['order_no'].astype(str).str.contains(search_order.strip(), case=False, na=False)]
            else:
                df_filtered = gc
                
            # 将最新的记录排在最前面
            df_filtered = df_filtered.sort_values("id", ascending=False)
            
            # 3. 动态展示结果与导出按钮
            if df_filtered.empty:
                st.warning("⚠️ 找不到该订单号的消耗记录！")
            else:
                st.dataframe(df_filtered, use_container_width=True, hide_index=True)
                
                # 导出 Excel 逻辑
                out_c = io.BytesIO()
                with pd.ExcelWriter(out_c, engine='openpyxl') as writer:
                    df_filtered.to_excel(writer, index=False, sheet_name='消耗记录')
                
                export_name = f"制衣厂消耗记录_{search_order.strip() if search_order else '全部'}_{datetime.date.today()}.xlsx"
                st.download_button(
                    label=f"📥 导出当前 {len(df_filtered)} 条检索结果为 Excel", 
                    data=out_c.getvalue(), 
                    file_name=export_name, 
                    type="primary",
                    use_container_width=True
                )
            
            st.divider()
            
            # 4. 撤销/删除模块
            st.markdown("##### 🗑️ 撤销消耗记录")
            c_sel, c_btn = st.columns([1, 2])
            with c_sel:
                cid_to_undo = st.selectbox("👉 选择要撤销的消耗记录 ID", df_filtered['id'].tolist(), key="sel_undo_consume")
            
            with c_btn:
                st.write("") # 占位对齐
                st.write("")
                if cid_to_undo:
                    # 权限校验：管理员可以撤销所有，业务员只能撤销自己填的
                    row_c = gc[gc['id']==cid_to_undo].iloc[0]
                    can_undo_c = (role == 'admin') or (row_c.get('operator') == user_name)
                    
                    if can_undo_c:
                        if st.button(f"🗑️ 彻底删除消耗单 #{cid_to_undo}", type="secondary"):
                            ok, msg = undo_consumption(cid_to_undo)
                            if ok: 
                                # 异步删除飞书中的对应记录
                                threading.Thread(target=delete_feishu_garment_consumption, args=(cid_to_undo,)).start()
                                st.success(msg)
                                time.sleep(1)
                                st.rerun()
                            else: 
                                st.error(msg)
                    else:
                        st.caption("🔒 您没有权限撤销其他人的记录")

     # ==========================================
    # 7. 跨境物料下单记录 (新增)
    # ==========================================
    with tab_cross:
        st.subheader("🌍 跨境物料下单记录")
        conn = get_db_conn()
        
        # 读取订单数据并关联物料信息
        try:
            df = pd.read_sql_query("""
                SELECT 
                    o.id AS 订单ID,
                    o.order_time AS 下单时间,
                    m.name_model AS 物料名称,
                    m.material_code AS 物料编码,
                    o.quantity AS 数量,
                    o.operator AS 操作人
                FROM crossborder_orders_v2 o
                JOIN crossborder_materials_v2 m ON o.material_id = m.id
                ORDER BY o.order_time DESC
            """, conn)
        except Exception as e:
            st.error(f"读取数据失败: {e}")
            df = pd.DataFrame()
        finally:
            conn.close()
        
        if df.empty:
            st.info("📭 暂无跨境物料下单记录")
        else:
            # ---------- 筛选功能 ----------
            st.markdown("##### 🔍 筛选与导出")
            # 日期范围筛选
            col1, col2 = st.columns(2)
            with col1:
                start_date = st.date_input("起始日期", value=df['下单时间'].min().date() if not df.empty else datetime.date.today())
            with col2:
                end_date = st.date_input("结束日期", value=df['下单时间'].max().date() if not df.empty else datetime.date.today())
            
            # 物料名称筛选
            all_materials = ["全部"] + sorted(df['物料名称'].unique().tolist())
            selected_material = st.selectbox("物料名称", all_materials)
            
            # 应用筛选
            df_filtered = df.copy()
            df_filtered['下单时间'] = pd.to_datetime(df_filtered['下单时间']).dt.date
            df_filtered = df_filtered[(df_filtered['下单时间'] >= start_date) & (df_filtered['下单时间'] <= end_date)]
            if selected_material != "全部":
                df_filtered = df_filtered[df_filtered['物料名称'] == selected_material]
            
            if df_filtered.empty:
                st.info("该筛选条件下无记录")
            else:
                st.success(f"共找到 {len(df_filtered)} 条记录")
                st.dataframe(df_filtered, use_container_width=True, hide_index=True)
                
                # ---------- 导出 Excel ----------
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_filtered.to_excel(writer, index=False, sheet_name='跨境物料订单')
                    worksheet = writer.sheets['跨境物料订单']
                    for column in worksheet.columns:
                        max_len = 0
                        col_letter = get_column_letter(column[0].column)
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_len:
                                    max_len = len(str(cell.value))
                            except:
                                pass
                        worksheet.column_dimensions[col_letter].width = min(max_len + 2, 50)
                st.download_button(
                    "📥 导出当前筛选结果 (Excel)",
                    data=output.getvalue(),
                    file_name=f"跨境物料订单_{start_date}_{end_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            # ---------- 撤销功能 ----------
            st.divider()
            st.markdown("##### 🗑️ 撤销订单")
            col_sel, col_btn = st.columns([1, 2])
            with col_sel:
                order_ids = df['订单ID'].tolist()
                selected_id = st.selectbox("选择要撤销的订单ID", order_ids, key="cb_undo_id")
            with col_btn:
                # 获取该订单的操作人（用于权限判断）
                row = df[df['订单ID'] == selected_id].iloc[0] if selected_id in df['订单ID'].values else None
                if row is not None:
                    can_undo = (role == 'admin') or (row['操作人'] == user_name)
                    if can_undo:
                        if st.button(f"🗑️ 撤销订单 #{selected_id}", type="secondary"):
                            ok, msg = undo_crossborder_order(selected_id)
                            if ok:
                                # 异步删除飞书中的对应记录
                                threading.Thread(target=delete_feishu_crossborder_order, args=(selected_id,)).start()
                                st.success(msg)
                                time.sleep(1)
                                st.rerun()
                            else:
                                st.error(msg)
                    else:
                        st.caption("🔒 您没有权限撤销此订单（仅管理员或本人可撤销）")

    # ========== 编辑采购合同对话框（必须放在所有标签页之后） ==========
    if 'editing_po_id' in st.session_state:
        po_id = st.session_state['editing_po_id']
        edit_purchase_order(po_id, role, user_name)

# --- 撤销采购合同记录 ---
def undo_purchase_order(record_id):
    conn = get_db_conn()
    c = conn.cursor()
    try:
        c.execute("DELETE FROM purchase_orders WHERE id=%s", (record_id,))
        conn.commit()  # 👈 MySQL 数据在此刻被成功抹除
        
        # 🌟 新增：启动后台线程，去飞书执行同步删除
        try:
            # 注意：带有参数的线程，必须用 args=(参数,) 的元组形式传递
            sync_thread = threading.Thread(target=delete_feishu_purchase_order, args=(record_id,))
            sync_thread.start()
        except Exception as e:
            pass  # 即使线程启动失败，也不影响前端提示撤销成功
            
        return True, f"✅ 采购合同 #{record_id} 已成功撤销并删除！飞书同步删除中..."
    except Exception as e:
        return False, f"撤销失败: {e}"
    finally:
        conn.close()

def undo_crossborder_order(order_id):
    """撤销跨境物料订单：恢复库存，删除订单记录"""
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        # 查询订单详情
        cursor.execute("SELECT material_id, quantity FROM crossborder_orders_v2 WHERE id=%s", (order_id,))
        row = cursor.fetchone()
        if not row:
            return False, "订单不存在"
        material_id, qty = row
        # 恢复库存
        cursor.execute("UPDATE crossborder_materials_v2 SET stock_quantity = stock_quantity + %s WHERE id=%s", (qty, material_id))
        # 删除订单
        cursor.execute("DELETE FROM crossborder_orders_v2 WHERE id=%s", (order_id,))
        conn.commit()
        return True, f"订单 #{order_id} 已撤销，库存已恢复"
    except Exception as e:
        conn.rollback()
        return False, f"撤销失败: {e}"
    finally:
        conn.close()

def undo_crossborder_order(order_id):
    """撤销跨境物料订单：恢复库存，删除订单记录"""
    conn = get_db_conn()
    cursor = conn.cursor()
    try:
        # 查询订单详情
        cursor.execute("SELECT material_id, quantity FROM crossborder_orders_v2 WHERE id=%s", (order_id,))
        row = cursor.fetchone()
        if not row:
            return False, "订单不存在"
        material_id, qty = row
        # 恢复库存
        cursor.execute("UPDATE crossborder_materials_v2 SET stock_quantity = stock_quantity + %s WHERE id=%s", (qty, material_id))
        # 删除订单
        cursor.execute("DELETE FROM crossborder_orders_v2 WHERE id=%s", (order_id,))
        conn.commit()
        return True, f"订单 #{order_id} 已撤销，库存已恢复"
    except Exception as e:
        conn.rollback()
        return False, f"撤销失败: {e}"
    finally:
        conn.close()