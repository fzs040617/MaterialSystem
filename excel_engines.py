import pandas as pd
import io
import json
import zipfile
import os
import datetime
import re
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as xlImage
import base64
from copy import copy

# 导入自定义模块
from config import STANDARD_MAP
from database import get_db_conn, load_data
from utils import adjust_column_width, convert_to_rmb_upper

# ==========================================
# 1. 包装袋与物料出货单引擎
# ==========================================

def generate_cart_excel_bytes_multi(order_cart, meta_common, operator):
    """生成包装袋出货单（支持单厂 Excel 或多厂 Zip）"""
    g_df = load_data("garment_factories").set_index('name')
    specs = load_data("bag_specs")
    img_map = {}
    if not specs.empty and 'image_path' in specs.columns:
        specs['key'] = specs['name'] + "|" + specs['size']
        img_map = dict(zip(specs['key'], specs['image_path']))

    groups = {}
    for item in order_cart:
        src = item['src_factory']
        if src not in groups: groups[src] = []
        groups[src].append(item)
    
    date_str = str(meta_common['date'])
    if len(groups) == 1:
        src = list(groups.keys())[0]
        data = _generate_single_excel_with_images(groups[src], src, g_df, img_map)
        return data, f"出货单_{src}_{date_str}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for src, items in groups.items():
                data = _generate_single_excel_with_images(items, src, g_df, img_map)
                zip_file.writestr(f"出货单_{src}_{date_str}.xlsx", data)
        return zip_buffer.getvalue(), f"批量出货单_{date_str}.zip", "application/zip"

def _generate_single_excel_with_images(items, src_factory, g_df, img_map):
    """【内部】生成带图片的单张出货单表格"""
    rows = []
    for item in items:
        full_contact_info = g_df.loc[item['dst_garment'], 'address'] if item['dst_garment'] in g_df.index else ''
        rows.append({
            "日期": item.get('date', datetime.date.today()),
            "源工厂": item['src_factory'],
            "销售平台": item['platform'],
            "商品名称": item['product_name'],
            "数量": item['qty'],
            "发往制衣厂": item['dst_garment'],
            "包装袋名称": item['bag_name'],
            "尺寸": item['bag_size'],
            "收货地址/联系方式": full_contact_info,
            "图片": "" 
        })
    df = pd.DataFrame(rows)
    df.insert(0, '序号', range(1, len(df) + 1))
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="出货单")
        ws = writer.sheets["出货单"]
        adjust_column_width(ws)
        img_col_idx = df.columns.get_loc("图片") + 1
        img_col_letter = get_column_letter(img_col_idx)
        for i in range(2, len(df) + 2):
            ws.row_dimensions[i].height = 60
        ws.column_dimensions[img_col_letter].width = 15
        for i, item in enumerate(items):
            key = item['bag_name'] + "|" + item['bag_size']
            path = img_map.get(key)
            if path and os.path.exists(path):
                try:
                    img = xlImage(path)
                    img.width = 80; img.height = 80
                    ws.add_image(img, f"{img_col_letter}{i+2}")
                except: pass
    return out.getvalue()

def generate_other_mat_excel_bytes_multi(order_cart, meta_common, operator):
    """生成其他物料出货单（多厂自动打包）"""
    g_df = load_data("garment_factories").set_index('name')
    groups = {}
    for item in order_cart:
        src = item['src_factory']
        if src not in groups: groups[src] = []
        groups[src].append(item)
    date_str = str(meta_common['date'])
    if len(groups) == 1:
        src = list(groups.keys())[0]
        data = _generate_single_other_mat_excel(groups[src], src, g_df)
        return data, f"其他物料出货单_{src}_{date_str}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
            for src, items in groups.items():
                data = _generate_single_other_mat_excel(items, src, g_df)
                zip_file.writestr(f"其他物料出货单_{src}_{date_str}.xlsx", data)
        return zip_buffer.getvalue(), f"批量其他物料出货单_{date_str}.zip", "application/zip"

def _generate_single_other_mat_excel(items, src_factory, g_df):
    """【内部】生成单张其他物料出货单"""
    rows = []
    for item in items:
        full_contact = g_df.loc[item['dst_garment'], 'address'] if item['dst_garment'] in g_df.index else ''
        rows.append({
            "日期": item.get('date', datetime.date.today()),
            "发货源工厂": item['src_factory'],
            "物料信息 (编码+名称)": item['material'],
            "数量": item['qty'],
            "发往制衣厂": item['dst_garment'],
            "收货地址/联系方式": full_contact
        })
    df = pd.DataFrame(rows)
    df.insert(0, '序号', range(1, len(df) + 1))
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="物料出货单")
        adjust_column_width(writer.sheets["物料出货单"])
    return out.getvalue()

# ==========================================
# 2. 辅料下单表引擎
# ==========================================

def accessory_text_width(value):
    """根据中英文混排估算 Excel 显示宽度"""
    if value is None:
        return 0

    text = str(value).strip()
    if not text or text.lower() == "nan":
        return 0

    width = 0
    for ch in text:
        # 中文、中文标点、全角字符按 2 个宽度估算
        width += 2 if ord(ch) > 255 else 1

    return width


def auto_fit_accessory_columns_compact(ws, main_max_col, min_width=5, padding=1):
    """
    辅料下单表主表列宽自动紧凑：
    - 只调整主表明细列，不让底部收件信息、图片、右侧执行标准撑宽
    - 根据标题和明细内容动态计算
    - 每列设置最大宽度，避免空白过大
    """
    # 按列名控制最大宽度，避免整体太松
    width_rules = {
        "69码": (5, 7),
        "商家编码": (8, 16),
        "货品编号": (8, 16),
        "货品名称": (10, 22),
        "规格名称": (8, 14),
        "规格码": (8, 14),
        "采购量": (6, 9),
        "吊牌采购量": (9, 12),
        "洗水唛采购量": (10, 14),
        "零售价格": (8, 10),
        "平台": (7, 12),
        "内部码": (8, 16),
    }

    max_row = ws.max_row

    for col_idx in range(1, main_max_col + 1):
        col_letter = get_column_letter(col_idx)
        header = ws.cell(row=1, column=col_idx).value
        header_text = str(header).strip() if header is not None else ""

        col_min_width, col_max_width = width_rules.get(header_text, (min_width, 18))

        max_len = accessory_text_width(header_text)

        # 只扫描主表明细区域：第1行标题 + 明细行
        for row_idx in range(2, max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value

            # 遇到底部汇总/收件信息后，不继续用这些长文本撑宽主表
            if col_idx == 1 and isinstance(value, str):
                if "收件信息：" in value or "洗水唛：" in value or "绿色吊牌" in value or "五张新吊牌" in value:
                    break

            max_len = max(max_len, accessory_text_width(value))

        adjusted_width = max(col_min_width, min(max_len + padding, col_max_width))
        ws.column_dimensions[col_letter].width = adjusted_width

def apply_accessory_font_name(ws, font_name="微软雅黑"):
    """统一辅料下单表所有文字字体，不改变原来的字号、颜色、加粗等格式"""
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue

            new_font = copy(cell.font)
            new_font.name = font_name
            cell.font = new_font

def generate_accessory_excel(uploaded_file, params):
    """清洗旺店通数据并生成辅料下单表"""
    file_ext = os.path.splitext(uploaded_file.name)[1].lower()
    df = pd.read_csv(uploaded_file) if file_ext == '.csv' else pd.read_excel(uploaded_file)
    df.columns = [str(col).strip() for col in df.columns]

    required_cols = ['商家编码', '货品编号', '货品名称', '规格名称', '规格码', '采购确认量', '零售价格', '平台']
    missing_cols = [col for col in required_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"原表单缺少必要列：{', '.join(missing_cols)}")

    # 输入表可以有“采购量”列，也可以没有；但输出表不再使用/显示“采购量”
    df = df[required_cols].copy()

    # 去掉旺店通原表自带的合计 / 总计 / 小计行，避免写入明细表，也避免数量重复统计
    summary_keywords = r'合计|总计|小计'
    summary_mask = pd.Series(False, index=df.index)
    for col in ['商家编码', '货品编号', '货品名称', '规格名称', '规格码', '平台']:
        summary_mask = summary_mask | df[col].fillna('').astype(str).str.strip().str.contains(
            summary_keywords,
            regex=True,
            na=False
        )
    df = df[~summary_mask].copy()

    # 只保留真实商品明细行
    product_key_mask = (
        df['货品编号'].fillna('').astype(str).str.strip().ne('') |
        df['货品名称'].fillna('').astype(str).str.strip().ne('') |
        df['规格码'].fillna('').astype(str).str.strip().ne('')
    )
    df = df[product_key_mask].copy()

    df['采购确认量'] = pd.to_numeric(df['采购确认量'], errors='coerce')
    df = df.dropna(subset=['采购确认量'])
    if df.empty:
        raise ValueError("原表单采购确认量无有效数据")

    df['吊牌采购量'] = df['采购确认量'].astype(int)
    df['货品名称'] = (
        df['货品名称']
        .fillna('')
        .astype(str)
        .str.replace(r'\(VIP\)|（VIP）', '', regex=True)
        .str.strip()
    )

    def clean_excel_value(value):
        if pd.isna(value):
            return ''
        text = str(value).strip()
        return '' if text.lower() == 'nan' else text

    def format_size_text(value):
        """规格码/号型格式：155/75(M+)，去掉 =\"...\" 或 =“...” 外壳"""
        text = clean_excel_value(value)
        if not text:
            return ''

        text = text.strip()

        # 去掉前面因为防公式注入加上的单引号
        if text.startswith("'"):
            text = text[1:].strip()

        # 处理 ="155/75(M+)"、=“155/75(M+)”、=155/75(M+)
        if text.startswith("="):
            text = text[1:].strip()

        # 去掉首尾各种引号
        text = text.strip()
        text = text.strip('"')
        text = text.strip("'")
        text = text.strip('“')
        text = text.strip('”')

        return text.strip()

    def format_price_with_rmb(value):
        """执行标准商品价格式：¥238，避免出现 238.0"""
        rmb_symbol = chr(165)  # 半角人民币/日元符号：¥

        text = clean_excel_value(value)
        if not text:
            return ''

        normalized = (
            text.replace('￥', '')
            .replace('¥', '')
            .replace(chr(65509), '')  # 去掉全角 ￥
            .replace(chr(165), '')    # 去掉半角 ¥
            .replace(',', '')
            .strip()
        )

        if not normalized or normalized.lower() == 'nan':
            return ''

        try:
            price_num = float(normalized)
        except (TypeError, ValueError):
            return f"{rmb_symbol}{normalized}"

        if abs(price_num - round(price_num)) < 1e-9:
            return f"{rmb_symbol}{int(round(price_num))}"

        price_text = f"{price_num:.10f}".rstrip('0').rstrip('.')
        return f"{rmb_symbol}{price_text}"

    def format_summary_qty(value):
        """底部汇总数量格式：367个，避免 367.0个"""
        try:
            num = float(value)
            if abs(num - round(num)) < 1e-9:
                return str(int(round(num)))
            return str(num).rstrip('0').rstrip('.')
        except Exception:
            return clean_excel_value(value)

    def get_accessory_summary_name(value):
        """底部汇总名称：显示完整辅料款式名称"""
        text = clean_excel_value(value)
        return text or "吊牌"

    def split_material_lines(value, max_lines=4):
        """洗水唛材质表：按行拆分，最多保留4行"""
        text = clean_excel_value(value)
        if not text:
            return []

        lines = []
        for line in str(text).splitlines():
            line = line.strip()
            if line:
                lines.append(line)

        return lines[:max_lines]

    def get_wash_label_image_path():
        """查找洗水唛有斜杠提示图片"""
        image_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "accessory_templates")
        filenames = [
            "洗水唛有斜杠.png",
            "洗水唛有斜杠.jpg",
            "洗水唛有斜杠.jpeg",
        ]

        for filename in filenames:
            image_path = os.path.join(image_dir, filename)
            if os.path.exists(image_path):
                return image_path

        return ""

    if params['has_wash'] == '有':
        wash_multiplier = 2 if params.get('is_two_pack') == '是' else 1
        df['洗水唛采购量'] = df['吊牌采购量'] * wash_multiplier
    else:
        df['洗水唛采购量'] = 0

    out_cols = {}
    missing_69_count = 0

    if params['has_69'] == '有':
        out_cols['69码'] = ''

    out_cols.update({
        '商家编码': df['商家编码'],
        '货品编号': df['货品编号'],
        '货品名称': df['货品名称'],
        '规格名称': df['规格名称'],
        '规格码': df['规格码'].map(format_size_text),
    })

    # 输出表不显示“采购量”列；输入表有/没有“采购量”都不影响生成
    out_cols.update({
        '洗水唛采购量': df['洗水唛采购量'],
        '吊牌采购量': df['吊牌采购量'],
        '零售价格': df['零售价格'],
        '平台': df['平台'],
        '内部码': params['internal_code']
    })

    def get_accessory_image_path(accessory_type):
        """按辅料款式匹配本地图片"""
        image_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "images", "accessory_templates")
        image_map = {
            "绿色吊牌+吊粒": ["绿色吊牌吊粒.png", "绿色吊牌吊粒.jpg", "绿色吊牌吊粒.jpeg"],
            "五张新吊牌+防伪带": ["五张新吊牌防伪带.png", "五张新吊牌防伪带.jpg", "五张新吊牌防伪带.jpeg"],
        }

        for filename in image_map.get(accessory_type, []):
            image_path = os.path.join(image_dir, filename)
            if os.path.exists(image_path):
                return image_path

        return ""

    df_out = pd.DataFrame(out_cols)
    df_out = df_out.replace({np.nan: '', pd.NA: '', None: ''})
    df_out = df_out.apply(
        lambda col: col.map(
            lambda x: '' if isinstance(x, str) and x.strip().lower() == 'nan'
            else (f"'{x}" if isinstance(x, str) and x.startswith('=') else x)
        )
        if col.dtype == object else col
    )

    def get_standard_info(product_name):
        rules = [
            ('舒适文胸', 'GB/T 8878-2023'),
            ('吊带背心', 'GB/T 8878-2023'),
            ('文胸', 'FZ/T 73012-2017'),
            ('背心', 'GB/T 8878-2023')
        ]
        for category, standard in rules:
            if category in product_name:
                style_name = product_name.replace(category, '', 1).strip()
                display_name = f"{category}（{style_name}）" if style_name else category
                return display_name, standard
        return product_name, ''

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_out.to_excel(writer, index=False, sheet_name='辅料下单表')
        ws = writer.sheets['辅料下单表']

        offset_col = ws.max_column + 2

        header_fill = PatternFill("solid", fgColor="9999FF")
        yellow_fill = PatternFill("solid", fgColor="FFFF00")
        header_font = Font(name="微软雅黑", bold=True)

        # 右侧执行标准区域字体：只放大右侧，不影响左边主表
        standard_title_font = Font(name="微软雅黑", bold=True, size=16)
        standard_font = Font(name="微软雅黑", bold=True, size=14)
        standard_red_font = Font(name="微软雅黑", bold=True, size=14, color="FF0000")

        for cell in ws[1]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

        first_row = df.iloc[0]
        template_name, std_val = get_standard_info(str(first_row.get('货品名称', '')))
        internal_code = clean_excel_value(params.get('internal_code', ''))

        material_text = clean_excel_value(params.get('material_text', ''))
        material_lines = []
        if params.get('has_wash') == '有':
            material_lines = split_material_lines(material_text)

        standard_start_row = 1

        # 右侧上方：洗水唛材质表，仅洗水唛=有且用户填写材质表时输出
        if material_lines:
            wash_start_row = 1

            cell = ws.cell(row=wash_start_row, column=offset_col, value="润微")
            cell.fill = yellow_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

            for idx, line in enumerate(material_lines, start=1):
                material_cell = ws.cell(row=wash_start_row + idx, column=offset_col, value=line)
                material_cell.alignment = Alignment(horizontal='left', vertical='center')

            warn_cell = ws.cell(row=wash_start_row, column=offset_col + 1, value="洗水唛有斜杠")
            warn_cell.fill = yellow_fill
            warn_cell.font = Font(name="微软雅黑", bold=True, color="FF0000", size=18)            
            warn_cell.alignment = Alignment(horizontal='center', vertical='center')

            wash_image_path = get_wash_label_image_path()
            if wash_image_path:
                try:
                    img = xlImage(wash_image_path)
                    img.width = 180
                    img.height = 90
                    ws.add_image(img, f"{get_column_letter(offset_col + 1)}{wash_start_row + 1}")

                    for row_idx in range(wash_start_row + 1, wash_start_row + 6):
                        ws.row_dimensions[row_idx].height = 20
                except Exception as e:
                    ws.cell(row=wash_start_row + 1, column=offset_col + 1, value=f"洗水唛图片插入失败：{e}")
            else:
                ws.cell(row=wash_start_row + 1, column=offset_col + 1, value="洗水唛标识图片未找到")

            # 给图片和材质表预留空间，执行标准区域整体下移
            standard_start_row = wash_start_row + max(len(material_lines) + 1, 6) + 1

        # 右侧执行标准区域：根据上方洗水唛材质表动态下移
        tag_data = [
            "润微",
            f"名称：{template_name}",
            f"货号：{clean_excel_value(first_row.get('货品编号', ''))}",
            f"号型：{format_size_text(first_row.get('规格码', ''))}",
            f"颜色：{clean_excel_value(first_row.get('规格名称', ''))}",
            f"商品价：{format_price_with_rmb(first_row.get('零售价格', ''))}",
            f"执行标准：{std_val}"
        ]

        for i, text in enumerate(tag_data):
            cell = ws.cell(row=standard_start_row + i, column=offset_col, value=text)
            cell.font = standard_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

            # “润微”标题更醒目
            if i == 0:
                cell.fill = yellow_fill
                cell.font = standard_title_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[standard_start_row + i].height = 24

            # 执行标准行黄底加粗
            elif i == len(tag_data) - 1:
                cell.fill = yellow_fill
                cell.font = standard_font
                ws.row_dimensions[standard_start_row + i].height = 22

            else:
                ws.row_dimensions[standard_start_row + i].height = 22

        # 仅保留执行标准右侧内部码，位置跟随执行标准区域动态下移
        if internal_code:
            internal_cell = ws.cell(
                row=standard_start_row + len(tag_data) - 1,
                column=offset_col + 1,
                value=f"内部码：{internal_code}"
            )
            internal_cell.font = standard_red_font
            internal_cell.alignment = Alignment(horizontal='left', vertical='center')

        summary_row = max(len(df_out) + 3, standard_start_row + len(tag_data) + 3)
        tag_qty = int(df['吊牌采购量'].sum())
        wash_qty = int(df['洗水唛采购量'].sum())
        accessory_type = clean_excel_value(params.get('accessory_type', ''))
        selected_factory_addr = clean_excel_value(params.get('selected_factory_addr', ''))
        accessory_summary_name = get_accessory_summary_name(accessory_type)

        current_row = summary_row

        # 底部汇总：按实际吊牌款式显示，不再单独写“吊牌总采购量 / 辅料款式 / 吊牌材质”
        ws.cell(
            row=current_row,
            column=1,
            value=f"{accessory_summary_name}：{format_summary_qty(tag_qty)}个"
        )
        current_row += 1

        ws.cell(
            row=current_row,
            column=1,
            value=f"洗水唛：{format_summary_qty(wash_qty)}个"
        )
        current_row += 1

        # 图片仍按辅料款式插入，但不再额外输出“辅料款式：xxx”
        if accessory_type:
            image_path = get_accessory_image_path(accessory_type)
            if image_path:
                try:
                    img = xlImage(image_path)
                    img.width = 360
                    img.height = 220
                    ws.add_image(img, f"A{current_row}")

                    # 给图片预留高度，避免图片压住收件信息
                    for row_idx in range(current_row, current_row + 11):
                        ws.row_dimensions[row_idx].height = 20

                    current_row += 12
                except Exception as e:
                    ws.cell(row=current_row, column=1, value=f"辅料图片插入失败：{e}")
                    current_row += 1

        ws.cell(row=current_row, column=1, value=f"收件信息：{selected_factory_addr}")

        # 主表列宽：根据实际字数自动紧凑调整
        auto_fit_accessory_columns_compact(ws, main_max_col=len(df_out.columns), min_width=5, padding=1)

        # 主表和右侧区域之间的空白列缩窄；右侧执行标准区域保持可读
        ws.column_dimensions[get_column_letter(offset_col - 1)].width = 3
        ws.column_dimensions[get_column_letter(offset_col)].width = 36
        ws.column_dimensions[get_column_letter(offset_col + 1)].width = 24

        apply_accessory_font_name(ws, "微软雅黑")

    return output.getvalue(), missing_69_count


# ==========================================
# 3. 采购合同与库存报表引擎
# ==========================================

def generate_rw_purchase_contract_excel(contract_no, factory_name, items):
    # ==========================================
    # 🌟 新增：从数据库中自动获取乙方（发货工厂）的地址、电话、负责人
    # ==========================================
    factory_address = "________________________________"
    factory_phone = "________________"
    factory_contact = "________________"
    
    try:
        conn = get_db_conn()
        cursor = conn.cursor()
        cursor.execute("SELECT address, contact, manager FROM packaging_factories WHERE name=%s", (factory_name,))
        res = cursor.fetchone()
        if res:
            # 如果数据库里有值就提取，如果是空(None)就依然保留下划线占位符
            factory_address = str(res[0]).strip() if res[0] else "________________________________"
            factory_phone = str(res[1]).strip() if res[1] else "________________"
            factory_contact = str(res[2]).strip() if res[2] else "________________"
        conn.close()
    except Exception as e:
        print(f"获取工厂信息失败: {e}")

    wb = Workbook()
    ws = wb.active
    ws.title = "原材料订购合同"
    
    # 字体与样式定义
    font_title1 = Font(name='微软雅黑', size=16, bold=True)
    font_title2 = Font(name='微软雅黑', size=18, bold=True)
    font_normal = Font(name='微软雅黑', size=11, bold=False)
    font_bold = Font(name='微软雅黑', size=11, bold=True)
    font_red = Font(name='微软雅黑', size=11, color="FF0000", bold=False)
    font_header = Font(name='微软雅黑', size=12, bold=True) 
    
    align_c = Alignment(horizontal='center', vertical='center')
    align_wrap_c = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_l = Alignment(horizontal='left', vertical='center')
    align_r = Alignment(horizontal='right', vertical='center')
    align_wrap_l = Alignment(horizontal='left', vertical='center', wrap_text=True) 
    border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    def make_merged_cell(r, start_c, end_c, text, font, align, border=None):
        ws.merge_cells(start_row=r, start_column=start_c, end_row=r, end_column=end_c)
        cell = ws.cell(row=r, column=start_c, value=text)
        cell.font = font
        cell.alignment = align
        if border:
            for c in range(start_c, end_c + 1):
                ws.cell(row=r, column=c).border = border

    def _text_display_width(value):
        """估算中英文混排文本宽度，用于合同明细行高自适应"""
        text = "" if value is None else str(value)
        width = 0
        for ch in text:
            width += 2 if ord(ch) > 255 else 1
        return width

    def _estimate_wrapped_lines(value, col_width):
        """根据列宽估算需要几行显示"""
        text_width = _text_display_width(value)
        if text_width <= 0:
            return 1

        safe_width = max(float(col_width or 10), 1.0)
        return max(1, int((text_width + safe_width - 1) // safe_width))


    # 1. 顶部公司名与合同标题
    make_merged_cell(1, 1, 13, "广州润微服装有限公司", font_title1, align_c)
    make_merged_cell(2, 1, 13, "原材料订购合同", font_title2, align_c)
    
    make_merged_cell(3, 1, 6, f"下单日期：{datetime.date.today().strftime('%Y-%m-%d')}", font_normal, align_l)
    make_merged_cell(3, 7, 13, f"合同编号：{contract_no}", font_bold, align_r)
    
    make_merged_cell(4, 1, 6, "甲  方：广州润微服装有限公司", font_normal, align_l)
    make_merged_cell(4, 7, 13, f"乙  方：{factory_name}", font_normal, align_l)
    
    # ==========================================
    # 🌟 修改点：将硬编码的下划线替换为我们上面查出来的变量
    # ==========================================
    make_merged_cell(5, 1, 6, "地  址：广东省广州市海珠区新港街道新港西路135号中山大学国家大学科技园B座14楼", font_normal, align_l)
    make_merged_cell(5, 7, 13, f"地  址：{factory_address}", font_normal, align_l)  # 👈 自动填入地址
    
    make_merged_cell(6, 1, 6, "电  话：13580417221", font_normal, align_l)
    make_merged_cell(6, 7, 13, f"电  话：{factory_phone}", font_normal, align_l)     # 👈 自动填入电话
    
    make_merged_cell(7, 1, 6, "联 系 人：廖洁仪", font_normal, align_l)
    make_merged_cell(7, 7, 13, f"联系人：{factory_contact}", font_normal, align_l)   # 👈 自动填入联系人/负责人
    
    make_merged_cell(8, 1, 13, "经双方友好协商，共同签订以下订购合同：", font_normal, align_l)
    make_merged_cell(9, 1, 13, "一、甲方在乙方订购以下物料：", font_bold, align_l)
    
    # 2. 动态表格绘制
    headers = ['物料编号', '物料名称', '材质', '颜色', '尺寸/规格', '收货标准', '数量', '单位', '单价含税含运费(元)', '金额(元)', '货期', '备注', '图片']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=h)
        cell.font = font_header
        cell.alignment = align_c
        cell.border = border_thin
        
    r = 11
    total_amount = 0.0
    
    # 合同明细区列宽，用于估算自动换行后的行高
    # 顺序对应：A-M列
    contract_col_widths = [12, 22, 12, 10, 12, 18, 8, 6, 18, 12, 12, 14, 11]

    for item in items:
        code = item.get('物料编号', '')
        name = item.get('物料名称', '')
        material = item.get('材质', '') 
        color = item.get('颜色/材质', item.get('颜色', ''))
        size = item.get('尺寸/规格', item.get('尺寸', ''))
        standard = item.get('收货标准', '')
        qty = item.get('数量', 0)
        unit = item.get('单位', 'Pcs')
        price = item.get('单价', 0.0)
        del_date = item.get('货期', '')
        remark = item.get('备注', '')
        
        if not name and not code: continue
        line_total = qty * price
        total_amount += line_total
        
        row_data = [code, name, material, color, size, standard, qty, unit, price, line_total, del_date, remark, ""]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = font_normal
            cell.border = border_thin
            cell.alignment = align_wrap_c
            
        # 根据本行长文本动态调整行高，重点解决“收货标准”内容显示不全
        row_line_count = max(
            _estimate_wrapped_lines(name, contract_col_widths[1]),
            _estimate_wrapped_lines(material, contract_col_widths[2]),
            _estimate_wrapped_lines(color, contract_col_widths[3]),
            _estimate_wrapped_lines(size, contract_col_widths[4]),
            _estimate_wrapped_lines(standard, contract_col_widths[5]),
            _estimate_wrapped_lines(remark, contract_col_widths[11]),
            1
        )

        row_height = 22 + (row_line_count - 1) * 16

        # 有图片时至少保留图片高度
        if str(item.get('图片', '')).strip():
            row_height = max(row_height, 70)

        # 防止极端长文本把行拉得过高
        ws.row_dimensions[r].height = min(max(row_height, 28), 180)

        # 处理并贴入图片
        img_str = str(item.get('图片', '')).strip()
        if img_str.startswith('data:image'):
            try:
                if "," in img_str:
                    header, encoded = img_str.split(",", 1)
                else:
                    encoded = img_str
                img_data = base64.b64decode(encoded)
                img = xlImage(io.BytesIO(img_data))
                img.width = 60
                img.height = 60
                ws.add_image(img, f"M{r}")
            except Exception as e:
                print(f"写入图片失败: {e}")
                ws.cell(row=r, column=13, value="图片解析失败")
        elif img_str.startswith('http'):
            ws.cell(row=r, column=13, value=img_str)

        r += 1
        
    # 3. 财务大小写合计行
    make_merged_cell(r, 1, 8, "合计总金额（小写）：", font_bold, align_r, border_thin)
    make_merged_cell(r, 9, 13, total_amount, font_bold, align_c, border_thin)
    r += 1
    
    make_merged_cell(r, 1, 8, "合计总金额（大写）：", font_bold, align_r, border_thin)
    upper_amount = convert_to_rmb_upper(total_amount)
    make_merged_cell(r, 9, 13, upper_amount, font_bold, align_c, border_thin)
    r += 1

    # 4. 律师过审版法律条款
    make_merged_cell(r, 1, 13, "二、交货数量（按合同数量交货）。", font_bold, align_l); r += 1
    make_merged_cell(r, 1, 13, "三、质量要求：", font_bold, align_l); r += 1
    
    clauses_3 = [
        ("1、责任归属：乙方对产品生产全流程质量、原材料真伪、工艺稳定性、出厂合格率负全部责任。甲方仅提供最终确认稿、材质/克重/尺寸/颜色定版样（以下统称“封板样”）。因甲方提供的封板样本身存在设计缺陷或违反法律法规所导致的质量问题，由甲方自行承担。", 35),
        ("2、乙方权限：在不改变设计稿、不降低封板样质量（包括外观、尺寸、材质、功能、寿命）的前提下，乙方可自主决定：工艺参数、同规格物料批次、生产排程、辅助加工方式（含粘合方式选择、盒子内部结构微调、抗压/密封增强措施）、检测方法、公差内微调，以及其他不改变设计稿、不降低封板样质量的同类生产调整事项。上述列举适用于所有包材类型，包括但不限于盒子、袋子、吊牌、防拆扣、贴纸、纸箱等。", 60),
        ("3、乙方无权决定的事项：乙方不得修改设计、不得降级换料、不得降低克重、不得改变印刷内容、不得降低任何功能指标（如抗压、密封、承重、粘性等），也不得进行其他任何导致产品质量下降的调整。", 35),
        ("4、质量标准：所有产品必须与甲方封板样/确认稿完全一致。封板样与本合同文字描述应保持一致。乙方在收到封板样及合同后3个工作日内应核对两者是否一致，如有不一致须书面通知甲方。若乙方未在该期限内提出异议，则视为封板样与合同描述一致。", 35),
        ("5、不一致的处理：如封样板与确认稿或合同不一致，则以实物封样板为准。", 35),
        ("6、公差：①克重公差 ±5%（按订单标称克重）；②尺寸公差 ±2mm（长 / 宽 / 高）；③材质按第一点表格中“材质”执行，不得使用回收料；④印刷无漏印 / 错色 / 重影，粘合无脱层 / 爆口”。⑤无特殊约定时，按行业通用验收标准AQL执行。", 35)
    ]
    for text, height in clauses_3:
        if text.startswith("5、"):
            make_merged_cell(r, 1, 13, text, font_red, align_wrap_l)
        else:
            make_merged_cell(r, 1, 13, text, font_normal, align_wrap_l)
        
        ws.row_dimensions[r].height = height
        r += 1

    # 5. 质量问题分级矩阵表格
    make_merged_cell(r, 1, 13, "四、质量问题分级处理", font_bold, align_l); r += 1
    
    make_merged_cell(r, 1, 2, "级别", font_bold, align_c, border_thin)
    make_merged_cell(r, 3, 7, "定义", font_bold, align_c, border_thin)
    make_merged_cell(r, 8, 13, "处理方式", font_bold, align_c, border_thin)
    r += 1
    
    matrix_data = [
        ("轻微", 
         "不影响正常使用、无明显客诉风险的质量瑕疵，如微小色差、少量毛边、局部轻微划痕等。甲方在检验或使用过程中挑出的此类问题产品。", 
         "甲方有权按实际挑出的不良品数量，从该批次货款中扣除相应价款（按单价计算），或要求乙方在下一批次中按同等数量免费补足。乙方无需承担额外违约金。", 65),
        ("中等", 
         "影响部分使用功能或外观，但未对甲方造成实际经济损失，且可通过返工或局部更换解决的质量问题。", 
         "乙方应在甲方通知后3个工作日内完成返工或更换，费用由乙方承担。返工或更换期间，原定货期不顺延，若因此造成甲方交货延迟，乙方按本合同第九条承担货期违约责任。", 65),
        ("严重", 
         "出现以下情形之一的：①材质错误（与封板样不符）；②偷工减料（如克重、层数、密度等明显低于封板样）；③存在安全隐患（如锐利边角、有毒有害物质超标）；④引发严重客户投诉或退货，对甲方造成实际损失。", 
         "甲方有权拒收或退回该批次所属订单的全部产品，乙方须立即根据甲方数量需求交付新的合格产品，乙方需按甲方实际损失的5至10倍支付违约金。同时甲方有权解除本合同，及取消后续的追加订单。", 65)
    ]
    for level, define, action, height in matrix_data:
        make_merged_cell(r, 1, 2, level, font_bold, align_c, border_thin)
        make_merged_cell(r, 3, 7, define, font_normal, align_wrap_l, border_thin)
        make_merged_cell(r, 8, 13, action, font_normal, align_wrap_l, border_thin)
        ws.row_dimensions[r].height = height
        r += 1

    # 6. 后续法律条款
    make_merged_cell(r, 1, 13, "五、质量责任持续时间：", font_bold, align_l); r += 1
    make_merged_cell(r, 1, 13, "1、乙方对本合同项下每批次产品的质量责任，自该批次甲方或指定验收方之日起，持续6个月；", font_normal, align_l); r += 1
    make_merged_cell(r, 1, 13, "2、在责任持续期内发现的任何质量问题（包括隐蔽瑕疵），乙方均应按本合同第四条处理。", font_normal, align_l); r += 1
    
    make_merged_cell(r, 1, 13, "六、如双方对质量问题存在争议时，以具备 CMA、CNAS 资质的包材类权威第三方检测机构的检测结论为最终依据。检测机构由双方共同指定，无法共同指定的，甲方有权向所在地的人民法院起诉处理。检测不合格则费用由乙方承担，检测合格则费用由甲方承担。", font_normal, align_wrap_l)
    ws.row_dimensions[r].height = 35; r += 1
    
    make_merged_cell(r, 1, 13, "七、付款方式：双月结或季结", font_bold, align_l); r += 1
    make_merged_cell(r, 1, 13, "八、交货地点：按甲方实际通知", font_bold, align_l); r += 1
    make_merged_cell(r, 1, 13, "九、货期违约责任：", font_bold, align_l); r += 1
    make_merged_cell(r, 1, 13, "2、未经甲方同意逾期交货的，每逾期一天，乙方按该批次货款的1%向甲方支付违约金，最高不超过该批次货款的10%；", font_normal, align_l); r += 1
    make_merged_cell(r, 1, 13, "3、逾期超过10天的，甲方有权解除合同，乙方应退还已付款项，并按该批次货款的30%支付违约金。", font_normal, align_l); r += 1
    
    make_merged_cell(r, 1, 13, "十、本合同如发生纠纷，当事人双方应及时协商解决，协商不成时，由甲方所在地的人民法院诉讼处理，法院费用、保全担保费、律师费等由败诉方承担。", font_normal, align_wrap_l)
    ws.row_dimensions[r].height = 35; r += 1
    
    make_merged_cell(r, 1, 13, "十一、本合同甲乙双方签字（盖章）生效，双方各执一份，传真件具有同等法律效力。", font_bold, align_l); r += 1
    
    r += 1
    
    # 7. 底部签名落款区
    make_merged_cell(r, 1, 6, "甲方：广州润微服装有限公司", font_bold, align_l)
    make_merged_cell(r, 7, 13, f"乙方：{factory_name}", font_bold, align_l)
    r += 1
    make_merged_cell(r, 1, 6, "委托代理人：", font_normal, align_l)
    make_merged_cell(r, 7, 13, "委托代理人：", font_normal, align_l)
    r += 1
    make_merged_cell(r, 1, 6, "签订日期：", font_normal, align_l)
    make_merged_cell(r, 7, 13, "签订日期：", font_normal, align_l)
    
    # 8. 严格设定列宽
    # F列“收货标准”适当加宽，避免长标准内容显示不全
    col_widths = [12, 22, 12, 10, 12, 18, 8, 6, 18, 12, 12, 14, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
        
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()

def generate_inventory_excel_with_images():
    """导出带图片的库存盘点表"""
    conn = get_db_conn()
    q = "SELECT i.*, b.image_path FROM inventory i LEFT JOIN bag_specs b ON i.bag_name=b.name AND i.bag_size=b.size"
    df = pd.read_sql_query(q, conn); conn.close()
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df[['factory_name','bag_name','bag_size','stock_quantity']].to_excel(writer, index=False, sheet_name='库存')
        ws = writer.sheets['库存']
        ws.column_dimensions['E'].width = 16
        for index, row in df.iterrows():
            ws.row_dimensions[index+2].height = 60
            if row['image_path'] and os.path.exists(row['image_path']):
                try:
                    img = xlImage(row['image_path']); img.width = 75; img.height = 75
                    ws.add_image(img, f"E{index+2}")
                except: pass
    return out.getvalue()

def generate_monthly_report_excel(year, month):
    """生成月度发货汇总多表页报表"""
    conn = get_db_conn()
    s, e = f"{year}-{month:02d}-01", f"{year+1}-01-01" if month==12 else f"{year}-{month+1:02d}-01"
    q = f"SELECT * FROM order_history WHERE order_date >= '{s}' AND order_date < '{e}'"
    df = pd.read_sql_query(q, conn); conn.close()
    if df.empty: return None
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='总流水', index=False)
        adjust_column_width(writer.sheets['总流水'])
    return out.getvalue()
